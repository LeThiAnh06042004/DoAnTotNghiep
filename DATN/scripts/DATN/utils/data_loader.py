import csv
import json
import openpyxl
import yaml
import xlrd
import sqlite3
from typing import List, Dict, Any, Optional, Union
from pathlib import Path


#kiểm tra file tồn tại
def check_file_exists(filepath: str) -> Path:
    file_path = Path(filepath)
    if not file_path.is_file():
        raise FileNotFoundError(f"File not found at: {file_path}") #nếu file ko tồn tại thì báo lỗi
    return file_path


def load_csv_data(file_path):
    try:
        try:
            with open(file_path, "r", encoding="utf-8") as csvfile:
                reader = csv.DictReader(csvfile) #đọc csv dùng csv.DictReader - mỗi hàng thành dict vs key là header
                return list(reader)
        except UnicodeDecodeError:
            # Nếu lỗi thì thử lại với encoding Windows
            with open(file_path, "r", encoding="utf-8-sig") as csvfile:
                reader = csv.DictReader(csvfile)
                return list(reader)
    except Exception as e:
        raise Exception(f"Lỗi đọc file CSV: {e}")


#khai báo hàm vs 2 tham số filepath và encoding
#kiểu trả về có thể là 1 ds các dict or 1 đối tg duy nhất
def load_json_data(filepath: str, encoding: str = 'utf-8') -> Union[List[Dict[str, Any]], Dict[str, Any]]:
    try:
        file_path = check_file_exists(filepath) #gọi hàm kiểm tra tồn tại

        if file_path.stat().st_size == 0:  #lấy kthuoc file tính = byte, 0 là file rỗng
            raise ValueError("File Json trống")

        with open(file_path, encoding=encoding) as jsonfile:
            data = json.load(jsonfile)  #đọc toàn bộ file và parse nd Json -> đối tg python
            #nếu file ko có j hợp lệ or list hoặc dict nhưng rỗng thì tb lỗi
            if data is None or (isinstance(data, (list, dict)) and not data):
                raise ValueError("File Json ko chứa dữ liệu")
            return data

    except Exception as e:
        raise Exception(f"Lỗi đọc file Json: {e}")


#khai báo hàm vs 2 tham số là filepath và tên sheet cần đọc -> ds các dict, mỗi dict là 1 dòng dl
def load_excel_data(filepath: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    try:
        file_path = check_file_exists(filepath)
        ext = file_path.suffix.lower() #suffix để xđ là xls hay xlsx

        # Nếu là file .xls → dùng xlrd
        if ext == ".xls":
            workbook = xlrd.open_workbook(file_path) #mở file
            #nếu có tên sheet thì đọc sheet đó, nếu ko thì đọc sheet đtiên
            sheet = workbook.sheet_by_name(sheet_name) if sheet_name else workbook.sheet_by_index(0)

            #đọc header
            headers = [sheet.cell_value(0, col).strip() for col in range(sheet.ncols)]
            if not headers or any(h == "" for h in headers):
                raise KeyError("Excel .xls không có header hợp lệ.")

            data = []
            for row in range(1, sheet.nrows): #bỏ qua header, duyệt ừ dòng t2
                #mỗi dòng là 1 dict
                row_data = {headers[col]: sheet.cell_value(row, col) for col in range(sheet.ncols)}
                if any(str(v).strip() for v in row_data.values()): #nếu dòng trống hoàn toàn thì bỏ qua
                    data.append(row_data)
            return data

        # Nếu là file .xlsx → dùng openpyxl
        workbook = openpyxl.load_workbook(file_path)
        if not workbook.sheetnames:
            raise ValueError("Không tìm thấy sheet trong file Excel")

        sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

        #lấy dòng 1, chuyển sang string, nếu cell -> none
        headers = [str(cell.value).strip() if cell.value else None for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        if not headers or any(h is None for h in headers):
            raise KeyError("Excel .xlsx không có header hợp lệ.")

        data = []
        for row in sheet.iter_rows(min_row=2):
            #ánh xạ ừng ô trong excel vào header t.ư, tránh lỗi nếu số cột dl > số header
            row_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
            if any(row_data.values()):
                data.append(row_data)

        if not data:
            raise ValueError("File Excel không có dữ liệu.")

        return data

    except Exception as e:
        raise Exception(f"Lỗi đọc file Excel: {e}")


#trả về 1 dict ỏ list các dict
def load_yaml_data(filepath: str, encoding: str = 'utf-8') -> Union[List[Dict[str, Any]], Dict[str, Any]]:
    try:
        file_path = check_file_exists(filepath)

        if file_path.stat().st_size == 0: #file rỗng
            raise ValueError("File YAML rỗng")

        with open(file_path, encoding=encoding) as yamlfile:
            data = yaml.safe_load(yamlfile) #tự động chuyển nd thành kiểu dl python tương ứng
            #nếu file ko có nd or list or dict rỗng thì báo lỗi
            if data is None or (isinstance(data, (list, dict)) and not data):
                raise ValueError("File YAML không có dữ liệu")
            return data

    except yaml.YAMLError as e:
        raise Exception(f"Định dạng không hợp lệ {filepath}: {e}")
    except Exception as e:
        raise Exception(f"Lỗi đọc file YAML: {e}")



def load_txt_data(file_path):
    cases = [] #ktao ds rỗng
    with open(file_path, "r", encoding="utf-8") as f:
        #đọc từng dòng, bỏ qua dòng rỗng or chỉ có khoảng trắng
        lines = [line.strip() for line in f if line.strip()]

    #nếu file chỉ có 1 dòng trống, nghĩa là ko có dl test thì trả về ds rỗng
    if len(lines) < 2:
        return cases

    headers = [h.strip() for h in lines[0].split(",")] #xđ dòng tiêu đề

    for line in lines[1:]:  #duyệt từ dòng t2 trở đi
        values = [v.strip() for v in line.split(",")] #tách dl từng cột theo dấu ,
        if len(values) != len(headers): #nếu số gt ko khớp vs số header thì bỏ qua dòng đó
            continue
        #gộp header và value thành dict
        case = dict(zip(headers, values))
        cases.append(case)

    return cases

#khai báo hàm vs 2 tham số là đg dẫn và tên bảng
def load_sqlite_data(db_path, table_name):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row #thay đổi factory, cho phép truy cập theo tên cột
    cursor = conn.cursor()
    cursor.execute(f"SELECT * FROM {table_name}")
    rows = cursor.fetchall() #lấy toàn bộ các dòng kq từ truy vấn
    conn.close()
    return [dict(row) for row in rows] #chuyển dl thành dict


